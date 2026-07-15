"""
phase16_exports.py — Phase 16 validation exports & report.

Generates, into phase16_exports/:
  Validation_Report.pdf / .xlsx / .csv
  Strategy_Scorecard.csv
  Trade_Review.csv
  AI_Recommendations.csv
and Phase16_Validation_Report.md (in the python dir).

All content derives from phase16_validation computations — honest markers,
no fabricated numbers. PAPER TRADING / RESEARCH ONLY.
"""

from __future__ import annotations

import csv
import os
from datetime import datetime, timezone
from typing import Any

import phase16_validation as v

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(BASE_DIR, "phase16_exports")
REPORT_MD = os.path.join(BASE_DIR, "Phase16_Validation_Report.md")

NA = "Insufficient Data"


def _fmt(x: Any) -> Any:
    return NA if x is None else x


def _rows_overview(data: dict) -> list[list]:
    keys = ["overall_validation_score", "maturity", "trading_days_completed",
            "completed_trades", "open_trades", "win_rate_pct", "profit_factor",
            "expectancy", "max_drawdown_pct", "sharpe_ratio", "avg_holding_days",
            "avg_risk_reward", "capital_start", "capital_now", "capital_growth_pct"]
    return [[k, _fmt(data.get(k))] for k in keys]


def _scorecard_rows(sc: dict) -> tuple[list[str], list[list]]:
    header = ["Strategy", "Trades", "Wins", "Losses", "WinRatePct", "ProfitFactor",
              "AvgReturnPct", "MaxDrawdownPct", "BestTrade", "WorstTrade", "Status",
              "Recommendation"]
    rows = []
    for s in sc["strategies"]:
        rows.append([s["strategy"], s["trades"], s["wins"], s["losses"],
                     _fmt(s["win_rate_pct"]), _fmt(s["profit_factor"]),
                     _fmt(s["avg_return_pct"]), _fmt(s["max_drawdown_pct"]),
                     f"{s['best_trade']['symbol']} ({s['best_trade']['pnl']})",
                     f"{s['worst_trade']['symbol']} ({s['worst_trade']['pnl']})",
                     s["status"], s["recommendation"]])
    if not rows:
        rows = [[NA] + [""] * 11]
    return header, rows


def _trade_review_rows(tr: dict) -> tuple[list[str], list[list]]:
    header = ["Symbol", "EntryTime", "ExitTime", "EntryPrice", "ExitPrice", "Qty",
              "PnL", "PnLPct", "Regime", "Confidence", "OpportunityScore", "Strategy",
              "ExitReason", "RiskPct", "RewardPct", "HoldingDays", "AIExplanation",
              "LessonsLearned", "WinningFactors", "LosingFactors"]
    rows = []
    for t in tr["trades"]:
        rows.append([t["symbol"], t["entry_time"], t["exit_time"], t["entry_price"],
                     t["exit_price"], t["quantity"], t["pnl"], t["pnl_pct"],
                     t["market_regime"], _fmt(t["confidence"]),
                     _fmt(t["opportunity_score"]), t["strategy"], t["exit_reason"],
                     _fmt(t["risk_pct"]), _fmt(t["reward_pct"]),
                     _fmt(t["holding_period_days"]), t["ai_explanation"],
                     "; ".join(t["lessons_learned"]), "; ".join(t["winning_factors"]),
                     "; ".join(t["losing_factors"])])
    if not rows:
        rows = [["No completed trades yet"] + [""] * 19]
    return header, rows


def _recs_rows(recs: dict) -> tuple[list[str], list[list]]:
    header = ["Type", "Target", "Detail", "Suggestion", "AutoApplied"]
    rows = [[r["type"], r["target"], r["detail"], r["suggestion"], "NO — never automatic"]
            for r in recs["recommendations"]]
    return header, rows


def _validation_sections() -> dict:
    return {
        "overview": v.validation_overview(),
        "scorecard": v.strategy_scorecard(),
        "confidence": v.confidence_validation(),
        "regimes": v.regime_validation(),
        "sectors": v.sector_validation(),
        "ai": v.ai_decision_validation(),
        "trade_review": v.trade_review(),
        "recommendations": v.improvement_recommendations(),
        "failures": v.failure_analysis(),
        "successes": v.success_analysis(),
        "timeline": v.validation_timeline(),
        "bugs": v.bug_detection(),
    }


def _write_csv(path: str, header: list[str], rows: list[list]):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def _validation_csv(path: str, d: dict):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Section", "Metric", "Value"])
        for k, val in _rows_overview(d["overview"]):
            w.writerow(["Overview", k, val])
        for b in d["confidence"]["bands"]:
            w.writerow(["Confidence", f"band {b['band']} win_rate_pct", _fmt(b["win_rate_pct"])])
        for r in d["regimes"]["regimes"]:
            w.writerow(["Regime", f"{r['regime']} win_rate_pct ({r['trades']} trades)",
                        _fmt(r["win_rate_pct"])])
        for s in d["sectors"]["sectors"]:
            if s["trades"]:
                w.writerow(["Sector", f"{s['sector']} win_rate_pct ({s['trades']} trades)",
                            _fmt(s["win_rate_pct"])])
        t = d["timeline"]
        for k in ("trading_days", "completed_trades", "confidence_calibration_pct",
                  "strategy_stability_pct", "production_readiness_pct", "maturity"):
            w.writerow(["Timeline", k, _fmt(t.get(k))])
        w.writerow(["Health", "bug_check_verdict", d["bugs"]["verdict"]])
        w.writerow(["Meta", "label", "PAPER / RESEARCH ONLY"])


def _validation_xlsx(path: str, d: dict) -> str | None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        return f"xlsx unavailable: {e}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    for row in _rows_overview(d["overview"]):
        ws.append(row)
    ws2 = wb.create_sheet("Strategy Scorecard")
    header, rows = _scorecard_rows(d["scorecard"])
    ws2.append(header)
    for r in rows:
        ws2.append(r)
    ws3 = wb.create_sheet("Confidence")
    ws3.append(["Band", "Trades", "Wins", "Losses", "WinRatePct", "AvgReturnPct", "AvgHoldingDays"])
    for b in d["confidence"]["bands"]:
        ws3.append([b["band"], b["trades"], b["wins"], b["losses"],
                    _fmt(b["win_rate_pct"]), _fmt(b["avg_return_pct"]),
                    _fmt(b["avg_holding_days"])])
    ws4 = wb.create_sheet("Trade Review")
    header, rows = _trade_review_rows(d["trade_review"])
    ws4.append(header)
    for r in rows:
        ws4.append(r)
    ws5 = wb.create_sheet("Recommendations")
    header, rows = _recs_rows(d["recommendations"])
    ws5.append(header)
    for r in rows:
        ws5.append(r)
    wb.save(path)
    return None


def _validation_pdf(path: str, d: dict) -> str | None:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas as pdfcanvas
    except Exception as e:
        return f"pdf unavailable: {e}"
    c = pdfcanvas.Canvas(path, pagesize=A4)
    width, height = A4
    y = height - 2 * cm

    def line(text: str, size: int = 10, bold: bool = False, gap: float = 0.55):
        nonlocal y
        if y < 2 * cm:
            c.showPage()
            y = height - 2 * cm
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(2 * cm, y, text[:110])
        y -= gap * cm

    line("Phase 16 — Paper Trading Validation Report", 16, True, 1.0)
    line(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} — PAPER / RESEARCH ONLY", 9)
    y -= 0.4 * cm
    line("Overview", 13, True, 0.8)
    for k, val in _rows_overview(d["overview"]):
        line(f"{k}: {val}")
    y -= 0.3 * cm
    line("Strategy Scorecard", 13, True, 0.8)
    for s in d["scorecard"]["strategies"] or []:
        line(f"{s['strategy']}: {s['trades']} trades, win rate {_fmt(s['win_rate_pct'])}%, "
             f"PF {_fmt(s['profit_factor'])}, status {s['status']}")
    if not d["scorecard"]["strategies"]:
        line(NA)
    y -= 0.3 * cm
    line("Validation Timeline", 13, True, 0.8)
    t = d["timeline"]
    line(f"Trading days {t['trading_days']}/{t['trading_days_goal']} — "
         f"completed trades {t['completed_trades']}/{t['completed_trades_goal']}")
    line(f"Confidence calibration: {_fmt(t['confidence_calibration_pct'])} — "
         f"strategy stability: {_fmt(t['strategy_stability_pct'])}")
    line(f"Production readiness: {t['production_readiness_pct']}% ({t['maturity']})", 11, True)
    y -= 0.3 * cm
    line("AI Recommendations (advisory only)", 13, True, 0.8)
    for r in d["recommendations"]["recommendations"]:
        line(f"[{r['type']}] {r['target']}: {r['detail']}")
        line(f"    Suggestion: {r['suggestion']}")
    y -= 0.3 * cm
    line("Health / Bug Detection", 13, True, 0.8)
    line(f"Verdict: {d['bugs']['verdict']} ({d['bugs']['checks_performed']} checks)")
    for i in d["bugs"]["issues"]:
        line(f"[{i['severity']}] {i['check']}: {i['detail']}")
    c.save()
    return None


def _report_md(d: dict) -> str:
    o, t = d["overview"], d["timeline"]
    strat_lines = "\n".join(
        f"| {s['strategy']} | {s['trades']} | {_fmt(s['win_rate_pct'])} | "
        f"{_fmt(s['profit_factor'])} | {s['status']} |"
        for s in d["scorecard"]["strategies"]) or f"| {NA} | | | | |"
    rec_lines = "\n".join(f"- **[{r['type']}] {r['target']}** — {r['detail']} "
                          f"_Suggestion: {r['suggestion']}_"
                          for r in d["recommendations"]["recommendations"])
    issue_lines = "\n".join(f"- [{i['severity']}] {i['check']}: {i['detail']}"
                            for i in d["bugs"]["issues"]) or "- No issues detected"
    return f"""# Phase 16 Validation Report — Paper Trading Validation & Strategy Proving

Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} — PAPER TRADING / RESEARCH ONLY.

## Features Completed (Phase 16)
- Paper Trading Validation dashboard (overall score, maturity, core statistics)
- Strategy scorecard with advisory statuses (never auto-disabled)
- Confidence band validation, market regime validation, sector validation
- AI decision validation (with honest gaps: HOLD correctness and false negatives
  are not trackable without outcome tracking of unexecuted recommendations)
- Trade review with lessons learned, winning and losing factors
- Weekly and monthly review generators
- AI improvement recommendations (advisory only)
- Failure and success analysis
- Validation timeline toward production readiness
- Automated bug detection health report
- Exports: PDF / XLSX / CSV / scorecard / trade review / recommendations

## Validation Summary
- Completed trades: **{o['completed_trades']}** (goal {t['completed_trades_goal']})
- Trading days: **{o['trading_days_completed']}** (goal {t['trading_days_goal']})
- Win rate: {_fmt(o['win_rate_pct'])}% · Profit factor: {_fmt(o['profit_factor'])} ·
  Expectancy: {_fmt(o['expectancy'])}
- Max drawdown: {_fmt(o['max_drawdown_pct'])}% · Sharpe: {_fmt(o['sharpe_ratio'])}
- Capital: ₹{o['capital_start']} → ₹{o['capital_now']} ({o['capital_growth_pct']}%)
- Confidence verdict: {d['confidence']['verdict']}
- {o.get('note') or 'Sample size is sufficient for preliminary conclusions.'}

## Strategy Ranking
| Strategy | Trades | Win Rate % | Profit Factor | Status |
|---|---|---|---|---|
{strat_lines}

## Trading Statistics
- Winning trades: {d['successes']['winning_trades']} · Losing trades: {d['failures']['losing_trades']}
- Common winning regimes: {', '.join(str(x) for x in d['successes']['common_regimes'])}
- Best confidence range (winners): {d['successes']['best_confidence_range']}

## AI Statistics
- BUY / WATCH / IGNORE recommendations: {d['ai']['buy_recommendations']} /
  {d['ai']['watch_recommendations']} / {d['ai']['ignore_recommendations']}
- Executed: {d['ai']['executed_recommendations']} ·
  Correct BUY %: {_fmt(d['ai']['correct_buy_pct'])} ·
  Correct EXIT %: {_fmt(d['ai']['correct_exit_pct'])}
- {d['ai']['note']}

## Risk Statistics
- Max drawdown: {_fmt(o['max_drawdown_pct'])}% · Sharpe: {_fmt(o['sharpe_ratio'])} ·
  Avg risk/reward: {_fmt(o['avg_risk_reward'])}

## Recommended Improvements (advisory only — never auto-applied)
{rec_lines}

## Known Issues / Health
- Bug detection verdict: **{d['bugs']['verdict']}** ({d['bugs']['checks_performed']} checks)
{issue_lines}
- Not checkable server-side: {', '.join(d['bugs']['not_checkable'])}

## Production Readiness Score
- **{t['production_readiness_pct']}% — {t['maturity']}**
- Confidence calibration: {_fmt(t['confidence_calibration_pct'])} ·
  Strategy stability: {_fmt(t['strategy_stability_pct'])}

_No new indicators or strategies were added in Phase 16. Live trading remains
impossible; broker execution, risk engine, AI decision logic and learning
governance are unchanged._
"""


def build_exports() -> dict:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    d = _validation_sections()
    warnings: list[str] = []

    _validation_csv(os.path.join(EXPORT_DIR, "Validation_Report.csv"), d)
    header, rows = _scorecard_rows(d["scorecard"])
    _write_csv(os.path.join(EXPORT_DIR, "Strategy_Scorecard.csv"), header, rows)
    header, rows = _trade_review_rows(d["trade_review"])
    _write_csv(os.path.join(EXPORT_DIR, "Trade_Review.csv"), header, rows)
    header, rows = _recs_rows(d["recommendations"])
    _write_csv(os.path.join(EXPORT_DIR, "AI_Recommendations.csv"), header, rows)

    err = _validation_xlsx(os.path.join(EXPORT_DIR, "Validation_Report.xlsx"), d)
    if err:
        warnings.append(err)
    err = _validation_pdf(os.path.join(EXPORT_DIR, "Validation_Report.pdf"), d)
    if err:
        warnings.append(err)

    with open(REPORT_MD, "w") as f:
        f.write(_report_md(d))

    files = sorted(os.listdir(EXPORT_DIR))
    return {
        "success": True,
        "export_dir": EXPORT_DIR,
        "files": files,
        "report_md": os.path.basename(REPORT_MD),
        "warnings": warnings,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "label": "PAPER / RESEARCH ONLY",
    }
