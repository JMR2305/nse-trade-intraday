"""
phase17_reports.py — Phase 17 automated QA reports.

Generates, into phase17_reports/:
  Validation_Report.pdf / .xlsx / .csv
  System_Health.json
  Release_Readiness.json
  Regression_Report.csv

All content derives from the most recent phase17_qa complete validation run —
honest markers, no fabricated numbers. PAPER TRADING / RESEARCH ONLY.
"""

from __future__ import annotations

import csv
import json
import os
from datetime import datetime, timezone
from typing import Any

import phase17_qa as q

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.path.join(BASE_DIR, "phase17_reports")

NA = "Insufficient Data"


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _all_check_rows(report: dict) -> list[list]:
    rows = []
    for key, sec in report.get("sections", {}).items():
        for c in sec.get("checks", []):
            rows.append([sec.get("section", key), c["check"], c["status"], c["detail"]])
    return rows


def _csv_report(path: str, report: dict) -> None:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Phase 17 Validation Report", report.get("generated_at", "")])
        w.writerow(["Version", report.get("release_version"), "Build",
                    report.get("build_number"), "Verdict", report.get("verdict"),
                    "Health Score", report.get("health_score")])
        w.writerow([])
        w.writerow(["Section", "Check", "Status", "Detail"])
        w.writerows(_all_check_rows(report))


def _xlsx_report(path: str, report: dict) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as e:
        return f"xlsx unavailable: {e}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    for row in [
        ["Phase 17 Validation Report", report.get("generated_at", "")],
        ["Version", report.get("release_version")],
        ["Build", report.get("build_number")],
        ["Verdict", report.get("verdict")],
        ["Health Score", report.get("health_score")],
        ["Total / Passed / Failed / Warnings",
         f"{report.get('total')} / {report.get('passed')} / "
         f"{report.get('failed')} / {report.get('warnings')}"],
        ["Duration (s)", report.get("duration_seconds")],
        ["Label", report.get("label")],
    ]:
        ws.append(row)
    ws["A1"].font = bold
    ws2 = wb.create_sheet("Checks")
    ws2.append(["Section", "Check", "Status", "Detail"])
    for c in ws2[1]:
        c.font = bold
    for row in _all_check_rows(report):
        ws2.append([str(x) for x in row])
    ws3 = wb.create_sheet("Checklist")
    ws3.append(["Item", "Status", "Detail"])
    for c in ws3[1]:
        c.font = bold
    for item in report.get("release_checklist", []):
        ws3.append([item["item"], item["status"], item["detail"]])
    wb.save(path)
    return None


def _pdf_report(path: str, report: dict) -> str | None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                        TableStyle)
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception as e:
        return f"pdf unavailable: {e}"
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    story = [
        Paragraph("Phase 17 Validation Report — Automated QA & Release Validation",
                  styles["Title"]),
        Paragraph(f"Generated {report.get('generated_at', '')} · "
                  f"Version {report.get('release_version')} · Build {report.get('build_number')} · "
                  f"PAPER TRADING / RESEARCH ONLY", styles["Normal"]),
        Spacer(1, 12),
        Paragraph(f"Verdict: <b>{report.get('verdict')}</b> · Health Score: "
                  f"<b>{report.get('health_score')}</b> / 100 · "
                  f"{report.get('passed')} passed, {report.get('failed')} failed, "
                  f"{report.get('warnings')} warning(s) of {report.get('total')} checks "
                  f"in {report.get('duration_seconds')}s", styles["Normal"]),
        Spacer(1, 12),
        Paragraph("Release Checklist", styles["Heading2"]),
    ]
    cl_rows = [["Item", "Status", "Detail"]] + [
        [i["item"], i["status"], i["detail"][:80]]
        for i in report.get("release_checklist", [])]
    t = Table(cl_rows, colWidths=[6 * cm, 2 * cm, 9 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
    ]))
    story += [t, Spacer(1, 12), Paragraph("All Checks", styles["Heading2"])]
    rows = [["Section", "Check", "Status"]] + [
        [r[0], r[1][:60], r[2]] for r in _all_check_rows(report)]
    t2 = Table(rows, colWidths=[5 * cm, 9 * cm, 2 * cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
    ]))
    story += [t2, Spacer(1, 12),
              Paragraph(report.get("score_note", ""), styles["Normal"])]
    doc.build(story)
    return None


def _regression_report(path: str, report: dict) -> None:
    """Compare the latest run's per-section scores to the previous run."""
    history = q._load_json(os.path.basename(q.HISTORY_PATH), [])
    prev = history[-2] if len(history) >= 2 else None
    rows = [["Metric", "Previous Run", "Current Run", "Delta", "Regression?"]]
    if prev is None:
        rows.append(["(first recorded run — no baseline to compare)", "", "", "", ""])
    else:
        for metric in ("passed", "failed", "warnings", "health_score"):
            a, b = prev.get(metric), report.get(metric)
            delta = (round(b - a, 1) if isinstance(a, (int, float))
                     and isinstance(b, (int, float)) else NA)
            worse = (metric == "failed" and isinstance(delta, (int, float)) and delta > 0) or \
                    (metric in ("passed", "health_score")
                     and isinstance(delta, (int, float)) and delta < 0)
            rows.append([metric, a, b, delta, "YES" if worse else "no"])
    with open(path, "w", newline="") as f:
        csv.writer(f).writerows(rows)


def build_reports(report: dict | None = None) -> dict:
    """Build all Phase 17 report files from the given (or last stored) run."""
    if report is None:
        report = q.last_run()
        if not report.get("available"):
            return {"success": False,
                    "error": "No validation run exists yet — run complete validation first."}
    os.makedirs(REPORT_DIR, exist_ok=True)
    warnings: list[str] = []

    _csv_report(os.path.join(REPORT_DIR, "Validation_Report.csv"), report)
    err = _xlsx_report(os.path.join(REPORT_DIR, "Validation_Report.xlsx"), report)
    if err:
        warnings.append(err)
    err = _pdf_report(os.path.join(REPORT_DIR, "Validation_Report.pdf"), report)
    if err:
        warnings.append(err)

    with open(os.path.join(REPORT_DIR, "System_Health.json"), "w") as f:
        json.dump({
            "generated_at": _now(), "label": report.get("label"),
            "health_score": report.get("health_score"),
            "verdict": report.get("verdict"),
            "section_scores": report.get("section_scores"),
            "totals": {k: report.get(k) for k in ("total", "passed", "failed", "warnings")},
        }, f, indent=2)

    with open(os.path.join(REPORT_DIR, "Release_Readiness.json"), "w") as f:
        json.dump({
            "generated_at": _now(), "label": report.get("label"),
            "release_version": report.get("release_version"),
            "build_number": report.get("build_number"),
            "production_ready": report.get("production_ready"),
            "release_checklist": report.get("release_checklist"),
        }, f, indent=2)

    _regression_report(os.path.join(REPORT_DIR, "Regression_Report.csv"), report)

    files = sorted(os.listdir(REPORT_DIR))
    return {"success": True, "generated_at": _now(), "label": report.get("label"),
            "report_dir": "phase17_reports", "files": files, "warnings": warnings}
